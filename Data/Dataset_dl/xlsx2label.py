import os
import sys
import time
import openpyxl
import re


def parse_time(time_str):
    match = re.match(r"((\d+)')?(\d+)''", time_str)
    if match:
        minutes = int(match.group(2)) if match.group(2) else 0
        seconds = int(match.group(3))
        return minutes * 60 + seconds
    return None


def process_data(data):
    time_str = data.split('-')[0]
    time = parse_time(time_str)

    if '비명' in data:
        emotion = 'scream'
    elif '당황' in data:
        emotion = 'dismay'
    else:
        emotion = 'null'

    if '패널' in data:
        emotion = 'panel_' + emotion
    if '한문철' in data:
        emotion = 'HMC_' + emotion

    return time, emotion


def extract_youtube_code(link):
    parts = link.split("/")
    code = parts[-1]
    code = code.split("?si=")[0]
    # print(str(code))
    return str(code)


def extract_time_emotion(row):
    orig_list = row[1:]
    mod_list = [str(item).replace('-', '\'\'-') if item is not None and '-' in str(item) else item for item in orig_list]
    flag_counter = 0
    for _ in range(len(mod_list)):
        if mod_list[flag_counter] == None:
            break
        else:
            time, emotion = process_data(mod_list[flag_counter])
            if time is None:
                print("---------------------------")
                print(mod_list[flag_counter])
                print("---------------------------")
                break
            # print(str(time-2) + ".000000\t" + str(time+4) + ".000000\t" + str(emotion) + "\t", end="\n")
            flag_counter += 1
            yield int(time), str(emotion)


def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
        data_list = []
        work_list = []

        code = extract_youtube_code(row[0].value)
        data_list.append(code)

        for cell in row:
            if cell.column < 6 and cell.value is not None:
                work_list.append(str(cell.value))
            elif cell.column > 6 and cell.value is not None:
                work_list.append(str(cell.value) + '한문철')
        
        time_emotion_generator = extract_time_emotion(work_list)
        for time, emotion in time_emotion_generator:
            # print(str(time-2) + ".000000\t" + str(time+4) + ".000000\t" + str(emotion) + "\t", end="\n")
            data_list.append(time)
            data_list.append(emotion)

            yield data_list
    workbook.close()
    return 0


def main(excel_file_path, output_folder):
    os.makedirs(output_folder, exist_ok=True)
    try:
        for labels in read_excel(excel_file_path):
            with open(output_folder + "\\" + labels[0] +'.txt', 'w') as file:
                for time, emotion in zip(labels[1::2], labels[2::2]):
                    file.write(str(time-2) + ".000000\t" + str(time+4) + ".000000\t" + str(emotion) + "\n")
    except Exception as e:
        print(e)
    print("DONE!")


if __name__ == "__main__":
    '''# Check input parameters
    if len(sys.argv) != 3:
        print("Check input parameters:")
        sys.exit(1)

    link_table = sys.argv[1]
    output_folder = sys.argv[2]'''

    currentdir = "C:\\Users\mosfet\source\\repos\AudioMNIST\datasetdl"
    link_table = currentdir + "\\adl_11111.xlsx"
    output_folder = currentdir + "\\Labels"

    main(link_table, output_folder)